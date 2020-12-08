#!/usr/bin/python

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Alignment, Side
import mysql.connector
from datetime import datetime, date, timedelta
import sys
import logger

from direcciones import encabezados_PON, encabezados_ONT, hojas_PON, hojas_ONT, excel_PON_semanal, excel_PON_mensual, excel_ONT_semanal, excel_ONT_mensual, hojas_pon, hojas_ont
from consultas import sql_ont_semanal, sql_ont_mensual, sql_pon_semanal, sql_pon_mensual
from conector import conector

def crear_hojas_ONT(workbook):

    # Creo Hojas
    subida_ont = workbook["Sheet"]
    subida_ont.title = hojas_ONT[0]
    bajada_ont = workbook.create_sheet(hojas_ONT[1])


def crear_hojas_PON(workbook):

    # Creo Hojas
    subida_pon = workbook["Sheet"]
    subida_pon.title = hojas_PON[0]
    bajada_pon = workbook.create_sheet(hojas_PON[1])
    subida_uplink = workbook.create_sheet(hojas_PON[2])
    bajada_uplink = workbook.create_sheet(hojas_PON[3])


def crear_encabezados_ONT(subida_ont,bajada_ont):
    #ENCABEZADOS
    # Subida Ont
    subida_ont["A1"] = encabezados_ONT[0]
    subida_ont["B1"] = encabezados_ONT[1]
    subida_ont["C1"] = encabezados_ONT[2]
    subida_ont["D1"] = encabezados_ONT[3]
    subida_ont["E1"] = encabezados_ONT[4]
    subida_ont["F1"] = encabezados_ONT[5]
    subida_ont["G1"] = encabezados_ONT[6]
    subida_ont["H1"] = encabezados_ONT[7]
    subida_ont["I1"] = encabezados_ONT[8]
    subida_ont["J1"] = encabezados_ONT[9]

    # Bajada Ont
    bajada_ont["A1"] = encabezados_ONT[0]
    bajada_ont["B1"] = encabezados_ONT[1]
    bajada_ont["C1"] = encabezados_ONT[2]
    bajada_ont["D1"] = encabezados_ONT[3]
    bajada_ont["E1"] = encabezados_ONT[4]
    bajada_ont["F1"] = encabezados_ONT[5]
    bajada_ont["G1"] = encabezados_ONT[6]
    bajada_ont["H1"] = encabezados_ONT[7]
    bajada_ont["I1"] = encabezados_ONT[8]
    bajada_ont["J1"] = encabezados_ONT[9]


def crear_encabezados_PON(subida_pon,bajada_pon,subida_uplink,bajada_uplink):
    #ENCABEZADOS
    # Subida Pon
    subida_pon["A1"] = encabezados_PON[0]
    subida_pon["B1"] = encabezados_PON[1]
    subida_pon["C1"] = encabezados_PON[2]
    subida_pon["D1"] = encabezados_PON[3]
    subida_pon["E1"] = encabezados_PON[4]
    subida_pon["F1"] = encabezados_PON[5]
    subida_pon["G1"] = encabezados_PON[6]
    subida_pon["H1"] = encabezados_PON[7]
    subida_pon["I1"] = encabezados_PON[8]
    subida_pon["J1"] = encabezados_PON[9]
    subida_pon["K1"] = encabezados_PON[10]
    subida_pon["L1"] = encabezados_PON[11]
    subida_pon["M1"] = encabezados_PON[12]
    # Bajada Pon
    bajada_pon["A1"] = encabezados_PON[0]
    bajada_pon["B1"] = encabezados_PON[1]
    bajada_pon["C1"] = encabezados_PON[2]
    bajada_pon["D1"] = encabezados_PON[3]
    bajada_pon["E1"] = encabezados_PON[4]
    bajada_pon["F1"] = encabezados_PON[5]
    bajada_pon["G1"] = encabezados_PON[6]
    bajada_pon["H1"] = encabezados_PON[7]
    bajada_pon["I1"] = encabezados_PON[8]
    bajada_pon["J1"] = encabezados_PON[9]
    bajada_pon["K1"] = encabezados_PON[10]
    bajada_pon["L1"] = encabezados_PON[11]
    bajada_pon["M1"] = encabezados_PON[12]

    # Subida Uplink
    subida_uplink["A1"] = encabezados_PON[0]
    subida_uplink["B1"] = encabezados_PON[1]
    subida_uplink["C1"] = encabezados_PON[2]
    subida_uplink["D1"] = encabezados_PON[3]
    subida_uplink["E1"] = encabezados_PON[4]
    subida_uplink["F1"] = encabezados_PON[5]
    subida_uplink["G1"] = encabezados_PON[6]
    subida_uplink["H1"] = encabezados_PON[7]
    subida_uplink["I1"] = encabezados_PON[8]

    # Bajada Uplink
    bajada_uplink["A1"] = encabezados_PON[0]
    bajada_uplink["B1"] = encabezados_PON[1]
    bajada_uplink["C1"] = encabezados_PON[2]
    bajada_uplink["D1"] = encabezados_PON[3]
    bajada_uplink["E1"] = encabezados_PON[4]
    bajada_uplink["F1"] = encabezados_PON[5]
    bajada_uplink["G1"] = encabezados_PON[6]
    bajada_uplink["H1"] = encabezados_PON[7]
    bajada_uplink["I1"] = encabezados_PON[8]


def apend_data_ONT(subida_ont,bajada_ont,periodo):
    if periodo == "semana":
        resultado = conector(sql_ont_semanal,"select","Select del reporte semanal de ONT")
    elif periodo == "mes":
        resultado = conector(sql_ont_mensual,"select","Select del reporte mensual de ONT")
    for dato in resultado:
        tipo = dato[1]
        nodo = dato[2]
        puerto = dato[3]
        etiqueta = dato[4]
        direccion = dato[5]
        hora = dato[6]
        fecha = dato[7]
        prom_hora_pico = dato[8]
        prom_picos_diarios = dato[9]
        pico = dato[10]
        if direccion == "RX":
            lista_append = [tipo,nodo,puerto,etiqueta,hora,fecha,pico,(pico*100)/10000,prom_hora_pico,prom_picos_diarios]
            subida_ont.append(lista_append)
        elif direccion == "TX":
            lista_append = [tipo,nodo,puerto,etiqueta,hora,fecha,pico,(pico*100)/10000,prom_hora_pico,prom_picos_diarios]
            bajada_ont.append(lista_append)


def apend_data_PON(subida_pon,bajada_pon,subida_uplink,bajada_uplink,periodo):
    if periodo == "semana":
        resultado = conector(sql_pon_semanal,"select","Select del reporte semanal de PON")
    elif periodo == "mes":
        resultado = conector(sql_pon_mensual,"select","Select del reporte mensual de PON")
    for dato in resultado:
        tipo = dato[1]
        nodo = dato[2]
        puerto = dato[3]
        direccion = dato[4]
        hora = dato[5]
        fecha = dato[6]
        prom_hora_pico = dato[7]
        prom_picos_diarios = dato[8]
        pico = dato[9]
        wf = dato[10]
        datos = dato[11]
        emp = dato[12]
        rbs = dato[13]
        if (direccion == "RX" and (puerto == "22/1" or puerto == "21/1")) or (direccion == "RX" and tipo == "MA5800" and (puerto == "9/0" or puerto == "10/0")):
            lista_append = [tipo,nodo,puerto,hora,fecha,pico,(pico*100)/10000,prom_hora_pico,prom_picos_diarios]
            bajada_uplink.append(lista_append)
        elif direccion == "TX" and (puerto == "22/1" or puerto == "21/1"):
            lista_append = [tipo,nodo,puerto,hora,fecha,pico,(pico*100)/10000,prom_hora_pico,prom_picos_diarios]
            subida_uplink.append(lista_append)
        elif direccion == "TX" and not (puerto == "21/2" or puerto == "21/3" or puerto == "21/4" or puerto == "22/2" or puerto == "22/3" or puerto == "22/4"):
            lista_append = [tipo,nodo,puerto,hora,fecha,pico,(pico*100)/2500,prom_hora_pico,prom_picos_diarios,wf,datos,emp,rbs]
            bajada_pon.append(lista_append)
        elif direccion == "RX" and not (puerto == "21/2" or puerto == "21/3" or puerto == "21/4" or puerto == "22/2" or puerto == "22/3" or puerto == "22/4"):
            lista_append = [tipo,nodo,puerto,hora,fecha,pico,(pico*100)/1250,prom_hora_pico,prom_picos_diarios,wf,datos,emp,rbs]
            subida_pon.append(lista_append)


def reportes_xlsx(tipo,periodo):
    workbook = Workbook()
    if tipo == "ONT" and periodo == "semana":
        logger.info("Comenzo a crearse el reporte semanal de ONT")
        crear_hojas_ONT(workbook)
        crear_encabezados_ONT(workbook[hojas_ONT[0]],workbook[hojas_ONT[1]])
        apend_data_ONT(workbook[hojas_ONT[0]],workbook[hojas_ONT[1]],periodo)
        workbook.save(filename=excel_ONT_semanal)
        logger.info("Se culmino la creacion del reporte semanal de ONT")

    elif tipo == "ONT" and periodo == "mes":
        logger.info("Comenzo a crearse el reporte mensual de ONT")
        crear_hojas_ONT(workbook)
        crear_encabezados_ONT(workbook[hojas_ONT[0]],workbook[hojas_ONT[1]])
        apend_data_ONT(workbook[hojas_ONT[0]],workbook[hojas_ONT[1]],periodo)
        workbook.save(filename=excel_ONT_mensual)
        logger.info("Se culmino la creacion del reporte mensual de ONT")

    elif  tipo == "PON" and periodo == "semana":
        logger.info("Comenzo a crearse el reporte semanal de PON")
        crear_hojas_PON(workbook)
        crear_encabezados_PON(workbook[hojas_PON[0]],workbook[hojas_PON[1]],workbook[hojas_PON[2]],workbook[hojas_PON[3]])
        apend_data_PON(workbook[hojas_PON[0]],workbook[hojas_PON[1]],workbook[hojas_PON[2]],workbook[hojas_PON[3]],periodo)
        workbook.save(filename=excel_PON_semanal)
        logger.info("Se culmino la creacion del reporte semanal de PON")

    elif tipo == "PON" and periodo == "mes":
        logger.info("Comenzo a crearse el reporte mensual de PON")
        crear_hojas_PON(workbook)
        crear_encabezados_PON(workbook[hojas_PON[0]],workbook[hojas_PON[1]],workbook[hojas_PON[2]],workbook[hojas_PON[3]])
        apend_data_PON(workbook[hojas_PON[0]],workbook[hojas_PON[1]],workbook[hojas_PON[2]],workbook[hojas_PON[3]],periodo)
        workbook.save(filename=excel_PON_mensual)
        logger.info("Se culmino la creacion del reporte mensual de PON")


#Codigo 0KM cheti cheti

def crear_hojas(workbook,titulos):
    contador = 0
    for titulo in titulos:
        if contador == 0:
            workbook["Sheet"].title = titulo.nombreHoja
            contador = contador + 1 
        else:
            workbook.create_sheet(titulo.nombreHoja)
    return workbook


def crear_encabezados(workbook,hojas):
    for hoja in hojas:
        for key in hoja.encabezados:
            workbook[hoja.nombreHoja][key] = hoja.encabezados[key]





def test():
    workbook = Workbook()
    crear_hojas(workbook,hojas_pon)
    crear_encabezados(workbook,hojas_pon)
    workbook.save(filename=excel_PON_semanal)


test()