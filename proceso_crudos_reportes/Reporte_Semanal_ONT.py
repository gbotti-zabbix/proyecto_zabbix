from openpyxl import Workbook
import mysql.connector
from datetime import datetime, date, timedelta

#Creo archivo y lo abro
filename = "Reporte" + " "+ str(date.today()) + ".xlsx"
workbook = Workbook()
encabezados = ["Modelo Nodo","Nodo","Slot/Puerto","Hora Pico","Fecha Pico","Pico","% Utilizacion","Prom. Hora Pico","Prom. Picos Diarios",""]


def crear_hojas(workbook):

    # Creo Hojas
    subida_pon = workbook["Sheet"]
    subida_pon.title = "Subida PON"
    bajada_pon = workbook.create_sheet("Bajada PON")
    subida_uplink = workbook.create_sheet("Subida Uplink")
    bajada_uplink = workbook.create_sheet("Bajada Uplink")

def crear_encabezados(subida_pon,bajada_pon,subida_uplink,bajada_uplink):
    #ENCABEZADOS
    # Subida Pon
    subida_pon["A1"] = encabezados[0]
    subida_pon["B1"] = encabezados[1]
    subida_pon["C1"] = encabezados[2]
    subida_pon["D1"] = encabezados[3]
    subida_pon["E1"] = encabezados[4]
    subida_pon["F1"] = encabezados[5]
    subida_pon["G1"] = encabezados[6]
    subida_pon["H1"] = encabezados[7]
    subida_pon["I1"] = encabezados[8]

    # Bajada Pon
    bajada_pon["A1"] = encabezados[0]
    bajada_pon["B1"] = encabezados[1]
    bajada_pon["C1"] = encabezados[2]
    bajada_pon["D1"] = encabezados[3]
    bajada_pon["E1"] = encabezados[4]
    bajada_pon["F1"] = encabezados[5]
    bajada_pon["G1"] = encabezados[6]
    bajada_pon["H1"] = encabezados[7]
    bajada_pon["I1"] = encabezados[8]

    # Subida Uplink
    subida_uplink["A1"] = encabezados[0]
    subida_uplink["B1"] = encabezados[1]
    subida_uplink["C1"] = encabezados[2]
    subida_uplink["D1"] = encabezados[3]
    subida_uplink["E1"] = encabezados[4]
    subida_uplink["F1"] = encabezados[5]
    subida_uplink["G1"] = encabezados[6]
    subida_uplink["H1"] = encabezados[7]
    subida_uplink["I1"] = encabezados[8]

    # Bajada Uplink
    bajada_uplink["A1"] = encabezados[0]
    bajada_uplink["B1"] = encabezados[1]
    bajada_uplink["C1"] = encabezados[2]
    bajada_uplink["D1"] = encabezados[3]
    bajada_uplink["E1"] = encabezados[4]
    bajada_uplink["F1"] = encabezados[5]
    bajada_uplink["G1"] = encabezados[6]
    bajada_uplink["H1"] = encabezados[7]
    bajada_uplink["I1"] = encabezados[8]

    return workbook

def apend_data(subida_pon,bajada_pon,subida_uplink,bajada_uplink):
    mydb = mysql.connector.connect(host="192.168.211.4",user="reportes",password="antel2020",database="reportes_zabbix")
    mycursor = mydb.cursor()
    mycursor.execute("SELECT * FROM reporte_semanal order by pico DESC")
    resultado = mycursor.fetchall()
    for dato in resultado:
        tipo = dato[1]
        nodo = dato[2]
        puerto = dato[3]
        direccion = dato[4]
        hora = dato[5]
        fecha = dato[6]
        pico = dato[9]
        prom_hora_pico = dato[7]
        prom_picos_diarios = dato[8]
        if (direccion == "RX" and (puerto == "22/1" or puerto == "21/1")) or (direccion == "RX" and tipo == "MA5800" and (puerto == "9/0" or puerto == "10/0")):
            lista_append = [tipo,nodo,puerto,hora,fecha,pico,(pico*100)/10000,prom_hora_pico,prom_picos_diarios]
            bajada_uplink.append(lista_append)
        elif direccion == "TX" and (puerto == "22/1" or puerto == "21/1"):
            lista_append = [tipo,nodo,puerto,hora,fecha,pico,(pico*100)/10000,prom_hora_pico,prom_picos_diarios]
            subida_uplink.append(lista_append)
        elif direccion == "TX" and not (puerto == "21/2" or puerto == "21/3" or puerto == "21/4" or puerto == "22/2" or puerto == "22/3" or puerto == "22/4"):
            lista_append = [tipo,nodo,puerto,hora,fecha,pico,(pico*100)/2500,prom_hora_pico,prom_picos_diarios]
            bajada_pon.append(lista_append)
        elif direccion == "RX" and not (puerto == "21/2" or puerto == "21/3" or puerto == "21/4" or puerto == "22/2" or puerto == "22/3" or puerto == "22/4"):
            lista_append = [tipo,nodo,puerto,hora,fecha,pico,(pico*100)/1250,prom_hora_pico,prom_picos_diarios]
            subida_pon.append(lista_append)
    return workbook

#Llamadas a la funcion y creacion de documento
crear_hojas(workbook)
crear_encabezados(workbook["Subida PON"],workbook["Bajada PON"],workbook["Subida Uplink"],workbook["Bajada Uplink"])
apend_data(workbook["Subida PON"],workbook["Bajada PON"],workbook["Subida Uplink"],workbook["Bajada Uplink"])
workbook.save(filename=filename)


def print_rows():
    for row in sheet.iter_rows(values_only=True):
        print(row)

