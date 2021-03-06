#!/usr/bin/python

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Alignment, Side
import mysql.connector
from datetime import datetime, date, timedelta
import sys

encabezados = ["Modelo Nodo","Nodo","Slot/Puerto","Hora Pico","Fecha Pico","Pico","% Utilizacion","Prom. Hora Pico","Prom. Picos Diarios","Total ONT","Servicios Datos","Servicios Empresariales","Empresariales de RBS"]

def crear_hojas(workbook):

    # Creo Hojas
    subida_pon = workbook["Sheet"]
    subida_pon.title = "Subida PON"
    bajada_pon = workbook.create_sheet("Bajada PON")
    subida_uplink = workbook.create_sheet("Subida Uplink")
    bajada_uplink = workbook.create_sheet("Bajada Uplink")

def crear_encabezados(subida_pon,bajada_pon,subida_uplink,bajada_uplink):
    #ENCABEZADOS
    # Subida Pon
    subida_pon["A1"] = encabezados[0]
    subida_pon["B1"] = encabezados[1]
    subida_pon["C1"] = encabezados[2]
    subida_pon["D1"] = encabezados[3]
    subida_pon["E1"] = encabezados[4]
    subida_pon["F1"] = encabezados[5]
    subida_pon["G1"] = encabezados[6]
    subida_pon["H1"] = encabezados[7]
    subida_pon["I1"] = encabezados[8]
    subida_pon["J1"] = encabezados[9]
    subida_pon["K1"] = encabezados[10]
    subida_pon["L1"] = encabezados[11]
    subida_pon["M1"] = encabezados[12]
    # Bajada Pon
    bajada_pon["A1"] = encabezados[0]
    bajada_pon["B1"] = encabezados[1]
    bajada_pon["C1"] = encabezados[2]
    bajada_pon["D1"] = encabezados[3]
    bajada_pon["E1"] = encabezados[4]
    bajada_pon["F1"] = encabezados[5]
    bajada_pon["G1"] = encabezados[6]
    bajada_pon["H1"] = encabezados[7]
    bajada_pon["I1"] = encabezados[8]
    bajada_pon["J1"] = encabezados[9]
    bajada_pon["K1"] = encabezados[10]
    bajada_pon["L1"] = encabezados[11]
    bajada_pon["M1"] = encabezados[12]

    # Subida Uplink
    subida_uplink["A1"] = encabezados[0]
    subida_uplink["B1"] = encabezados[1]
    subida_uplink["C1"] = encabezados[2]
    subida_uplink["D1"] = encabezados[3]
    subida_uplink["E1"] = encabezados[4]
    subida_uplink["F1"] = encabezados[5]
    subida_uplink["G1"] = encabezados[6]
    subida_uplink["H1"] = encabezados[7]
    subida_uplink["I1"] = encabezados[8]

    # Bajada Uplink
    bajada_uplink["A1"] = encabezados[0]
    bajada_uplink["B1"] = encabezados[1]
    bajada_uplink["C1"] = encabezados[2]
    bajada_uplink["D1"] = encabezados[3]
    bajada_uplink["E1"] = encabezados[4]
    bajada_uplink["F1"] = encabezados[5]
    bajada_uplink["G1"] = encabezados[6]
    bajada_uplink["H1"] = encabezados[7]
    bajada_uplink["I1"] = encabezados[8]
    
    return workbook

def apend_data(subida_pon,bajada_pon,subida_uplink,bajada_uplink,periodo):
    sql = ""
    mydb = mysql.connector.connect(host="localhost",user="reportes",password="antel2020",database="reportes_zabbix")
    mycursor = mydb.cursor()
    if periodo == "semana":
        sql = "SELECT * FROM reporte_semanal order by pico DESC"
    elif periodo == "mes":
        sql = "SELECT * FROM reporte_mensual order by pico DESC"
    mycursor.execute(sql)
    resultado = mycursor.fetchall()
    for dato in resultado:
        tipo = dato[1]
        nodo = dato[2]
        puerto = dato[3]
        direccion = dato[4]
        hora = dato[5]
        fecha = dato[6]
        prom_hora_pico = dato[7]
        prom_picos_diarios = dato[8]
        pico = dato[9]
        wf = dato[10]
        datos = dato[11]
        emp = dato[12]
        rbs = dato[13]
        if (direccion == "RX" and (puerto == "22/1" or puerto == "21/1")) or (direccion == "RX" and tipo == "MA5800" and (puerto == "9/0" or puerto == "10/0")):
            lista_append = [tipo,nodo,puerto,hora,fecha,pico,(pico*100)/10000,prom_hora_pico,prom_picos_diarios]
            bajada_uplink.append(lista_append)
        elif direccion == "TX" and (puerto == "22/1" or puerto == "21/1"):
            lista_append = [tipo,nodo,puerto,hora,fecha,pico,(pico*100)/10000,prom_hora_pico,prom_picos_diarios]
            subida_uplink.append(lista_append)
        elif direccion == "TX" and not (puerto == "21/2" or puerto == "21/3" or puerto == "21/4" or puerto == "22/2" or puerto == "22/3" or puerto == "22/4"):
            lista_append = [tipo,nodo,puerto,hora,fecha,pico,(pico*100)/2500,prom_hora_pico,prom_picos_diarios,wf,datos,emp,rbs]
            bajada_pon.append(lista_append)
        elif direccion == "RX" and not (puerto == "21/2" or puerto == "21/3" or puerto == "21/4" or puerto == "22/2" or puerto == "22/3" or puerto == "22/4"):
            lista_append = [tipo,nodo,puerto,hora,fecha,pico,(pico*100)/1250,prom_hora_pico,prom_picos_diarios,wf,datos,emp,rbs]
            subida_pon.append(lista_append)
    return workbook

def llamadas():
    crear_hojas(workbook)
    crear_encabezados(workbook["Subida PON"],workbook["Bajada PON"],workbook["Subida Uplink"],workbook["Bajada Uplink"])
    apend_data(workbook["Subida PON"],workbook["Bajada PON"],workbook["Subida Uplink"],workbook["Bajada Uplink"])
    workbook.save(filename=filename)

#MENU
if sys.argv[1] == "semanal":
    #Creo archivo y lo abro
    filename = "/var/lib/reportes-zabbix/reportes_semanales/Reporte_Semanal_" + str(date.today()) + ".xlsx"
    workbook = Workbook()
    #
    crear_hojas(workbook)
    crear_encabezados(workbook["Subida PON"],workbook["Bajada PON"],workbook["Subida Uplink"],workbook["Bajada Uplink"])
    apend_data(workbook["Subida PON"],workbook["Bajada PON"],workbook["Subida Uplink"],workbook["Bajada Uplink"],"semana")
    workbook.save(filename=filename)
elif sys.argv[1] == "mensual":
    #Creo archivo y lo abro
    filename = "/var/lib/reportes-zabbix/reportes_mensuales/Reporte_Mensual_" + str(date.today()) + ".xlsx"
    workbook = Workbook()
    #
    crear_hojas(workbook)
    crear_encabezados(workbook["Subida PON"],workbook["Bajada PON"],workbook["Subida Uplink"],workbook["Bajada Uplink"])
    apend_data(workbook["Subida PON"],workbook["Bajada PON"],workbook["Subida Uplink"],workbook["Bajada Uplink"],"mes")
    workbook.save(filename=filename)
