#!/usr/bin/python

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Alignment, Side
import mysql.connector
from datetime import datetime, date, timedelta
import sys

encabezados = ["Modelo Nodo","Nodo","Slot/Puerto","Etiqueta","Hora Pico","Fecha Pico","Pico","% Utilizacion","Prom. Hora Pico","Prom. Picos Diarios"]

def crear_hojas(workbook):

    # Creo Hojas
    subida_ont = workbook["Sheet"]
    subida_ont.title = "Subida ONT"
    bajada_ont = workbook.create_sheet("Bajada ONT")

def crear_encabezados(subida_ont,bajada_ont):
    #ENCABEZADOS
    # Subida Ont
    subida_ont["A1"] = encabezados[0]
    subida_ont["B1"] = encabezados[1]
    subida_ont["C1"] = encabezados[2]
    subida_ont["D1"] = encabezados[3]
    subida_ont["E1"] = encabezados[4]
    subida_ont["F1"] = encabezados[5]
    subida_ont["G1"] = encabezados[6]
    subida_ont["H1"] = encabezados[7]
    subida_ont["I1"] = encabezados[8]
    subida_ont["J1"] = encabezados[9]

    # Bajada Ont
    bajada_ont["A1"] = encabezados[0]
    bajada_ont["B1"] = encabezados[1]
    bajada_ont["C1"] = encabezados[2]
    bajada_ont["D1"] = encabezados[3]
    bajada_ont["E1"] = encabezados[4]
    bajada_ont["F1"] = encabezados[5]
    bajada_ont["G1"] = encabezados[6]
    bajada_ont["H1"] = encabezados[7]
    bajada_ont["I1"] = encabezados[8]
    bajada_ont["J1"] = encabezados[9]

    return workbook

def apend_data(subida_ont,bajada_ont,periodo):
    sql = ""
    mydb = mysql.connector.connect(host="localhost",user="reportes",password="antel2020",database="reportes_zabbix")
    mycursor = mydb.cursor()
    if periodo == "semana":
        sql = "SELECT * FROM reporte_semanal_ont order by pico DESC"
    elif periodo == "mes":
        sql = "SELECT * FROM reporte_mensual_ont order by pico DESC"
    mycursor.execute(sql)
    resultado = mycursor.fetchall()
    for dato in resultado:
        tipo = dato[1]
        nodo = dato[2]
        puerto = dato[3]
        etiqueta = dato[4]
        direccion = dato[5]
        hora = dato[6]
        fecha = dato[7]
        prom_hora_pico = dato[8]
        prom_picos_diarios = dato[9]
        pico = dato[10]
        if direccion == "RX":
            lista_append = [tipo,nodo,puerto,etiqueta,hora,fecha,pico,(pico*100)/10000,prom_hora_pico,prom_picos_diarios]
            subida_ont.append(lista_append)
        elif direccion == "TX":
            lista_append = [tipo,nodo,puerto,etiqueta,hora,fecha,pico,(pico*100)/10000,prom_hora_pico,prom_picos_diarios]
            bajada_ont.append(lista_append)
    return workbook

def llamadas(workbook,filename,periodo):
    crear_hojas(workbook)
    crear_encabezados(workbook["Subida ONT"],workbook["Bajada ONT"])
    apend_data(workbook["Subida ONT"],workbook["Bajada ONT"],periodo)
    workbook.save(filename=filename)

#MENU
if sys.argv[1] == "semanal":
    #Creo archivo y lo abro
    filename = "/var/lib/reportes-zabbix/reportes_semanales/Reporte_Semanal_" + str(date.today()) + "_ONT.xlsx"
    workbook = Workbook()
    #
    llamadas(workbook,filename,"semana")

elif sys.argv[1] == "mensual":
    #Creo archivo y lo abro
    filename = "/var/lib/reportes-zabbix/reportes_mensuales/Reporte_Mensual_" + str(date.today()) + "_ONT.xlsx"
    workbook = Workbook()
    #
    llamadas(workbook,filename,"mes")
